# -*- encoding: utf8 -*-

from threading import Thread
from openpyxl import load_workbook
from openpyxl import Workbook
from konlpy.tag import Kkma

scoredWords = dict()
kkma = Kkma()
movieCode = 0

def parseSentence(score, sentence) :
	for word in kkma.pos(sentence) :
		if word[1] in ('MA', 'MAG', 'MAC', 'MD', 'MDT', 'NN', 'NNG', 'NNP', 'NP', 'OH', 'OL', 'VA', 'VXA', 'VV', 'VXV', 'VX', 'XP', 'XPN', 'XPV', 'XSA', 'XSN', 'XSV', 'XR', 'UN'):
			if word[0] in scoredWords :
				scoredWords[word[0]].append(score)
			else :
				scoredWords[word[0]] = [score,]
	
def parseExcel(excel) :
	print("start parsing")

	sheet = excel.active
	maxRow = sheet.max_row

	print("max row :", maxRow)

	for sheetRow in range(2, maxRow) :
		print("parsing row num :", sheetRow)

		score = int(sheet['A' + str(sheetRow)].value)
		sentence = sheet['B' + str(sheetRow)].value

		parseSentence(score, sentence)

	print("end start parsing")
	saveToExcel()

def saveToExcel() :
	print("start saving to excel")

	wb = Workbook()
	ws = wb.active

	ws['A1'] = "단어"
	ws['B1'] = "평점"
	ws['C1'] = "빈도"

	wordKeys = list(scoredWords.keys())
	
	scoreAverage = 0

	for i in range(len(wordKeys)) :
		scoreFreq = len(scoredWords[wordKeys[i]])
		scoreSum = 0
		for score in scoredWords[wordKeys[i]] :
			scoreSum = scoreSum + score

		score = scoreSum / scoreFreq
		scoreAverage = scoreAverage + score
		
		ws['A' + str(i + 2)] = wordKeys[i]					# word
		ws['B' + str(i + 2)] = score						# score(avr)
		ws['C' + str(i + 2)] = scoreFreq					# freq
	
	scoreAverage = scoreAverage / len(wordKeys)
	ws['D1'] = scoreAverage
			
	wb.save("./" + str(movieCode) + "/word.xlsx")

	print("end saving to excel")

# thread를 정말 사용하고 싶었으나 konlpy가 thread-safe하지 않은 것 같음.

def main() :
	global movieCode
	movieCode = int(input("movie code : "))

	excel = load_workbook("./" + str(movieCode) + "/sentence_learn.xlsx")
	parseExcel(excel)


if __name__ == '__main__' :
	main()