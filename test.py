# -*- encoding: utf8 -*-

from urllib.request import urlopen
from bs4 import BeautifulSoup
from threading import Thread
from openpyxl import load_workbook
from openpyxl import Workbook
from konlpy.tag import Kkma

scoredWords = dict()
kkma = Kkma()
movieCode = 0
scoreAverage = 0

def scoreSentence(sentence) :
	scoreList = list()

	for word in kkma.pos(sentence) :
		if word[1] in ('MA', 'MAG', 'MAC', 'MD', 'MDT', 'NN', 'NNG', 'NNP', 'NP', 'OH', 'OL', 'VA', 'VXA', 'VV', 'VXV', 'VX', 'XP', 'XPN', 'XPV', 'XSA', 'XSN', 'XSV', 'XR', 'UN'):
			if word[0] in scoredWords and scoredWords[word[0]][1] >= 10 :
				scoreList.append(scoredWords[word[0]][0])
	
	scoreSum = 0

	for score in scoreList :
		scoreSum = scoreSum + score
	
	if len(scoreList) == 0 :
		return 0

	return scoreSum / len(scoreList)

def parseWordExcel() :
	global scoreAverage

	print("start parsing word excel")
	excel = load_workbook("./" + str(movieCode) + "/word.xlsx")

	sheet = excel.active
	maxRow = sheet.max_row

	scoreAverage = int(sheet['D1'].value)

	print("max row :", maxRow)

	for i in range(2, maxRow):
		word  = sheet['A' + str(i)].value
		score = float(sheet['B' + str(i)].value)
		freq  = float(sheet['C' + str(i)].value)

		scoredWords[word] = (score, freq)
	print("end parsing word excel")

def parseSentenceExcel() :
	print("start parsing sentence excel")
	excel = load_workbook("./" + str(movieCode) + "/sentence_test.xlsx")

	sheet = excel.active
	maxRow = sheet.max_row

	print("max row :", maxRow)

	testData = list()

	for i in range(2, maxRow):
		score    = int(sheet['A' + str(i)].value)
		sentence = sheet['B' + str(i)].value

		testData.append((score, sentence))

	print("end parsing sentence excel")
	return testData

def startTest(testData) :
	print("start test")
	positiveTestCount = 0
	positiveCorrectCount = 0
	negativeTestCount = 0
	negativeCorrectCount = 0
	for data in testData :
		score = scoreSentence(data[1])
		if data[0] >= 6 :
			positiveTestCount = positiveTestCount + 1
			if score >= scoreAverage :
				positiveCorrectCount = positiveCorrectCount + 1
				print("positive correct")
			else :
				print("positive wrong")
		else :
			negativeTestCount = negativeTestCount + 1
			if score < scoreAverage :
				negativeCorrectCount = negativeCorrectCount + 1
				print("negative correct")
			else :
				print("negative wrong")


	print("end test")
	return (positiveTestCount, negativeTestCount, positiveCorrectCount, negativeCorrectCount)

def printResult(testResult) :
	print("start printing result")
	print("")
	print("전체 횟수 : ", testResult[0] + testResult[1])
	print("정답 횟수 : ", testResult[2] + testResult[3])
	print("전체 정답률 : ", ((testResult[2] + testResult[3]) / (testResult[0] + testResult[1])))
	print("")
	print("긍정 횟수 : ", testResult[0])
	print("긍정 정답 횟수 : ", testResult[2])
	print("긍정 정답률 : ", (testResult[2] / testResult[0]))
	print("")
	print("부정 횟수 : ", testResult[1])
	print("부정 정답 횟수 : ", testResult[3])
	print("부정 정답률 : ", (testResult[3] / testResult[1]))
	print("")
	print("end printing result")

def main() :
	global movieCode
	movieCode = int(input("movie code : "))

	parseWordExcel()
	testData = parseSentenceExcel()
	testResult = startTest(testData)
	printResult(testResult)
	
	print("finished")

if __name__ == '__main__' :
	main()
